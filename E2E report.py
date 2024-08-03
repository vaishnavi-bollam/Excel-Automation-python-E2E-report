import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
import os
from tkinter import filedialog

print("Select input filepath")
input_file_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("Excel Files", "*.xlsx")])
if not input_file_path:
    print("Input file not selected. Exiting...")
    exit()

print("Create output file with desired name")
output_file_path = filedialog.asksaveasfilename(title="Save Output File", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
if not output_file_path:
    print("No location selected to save the output file. Exiting...")
    exit()

print("Select HCL queues filepath")
hcl_df_path = filedialog.askopenfilename(title="Select HCL Queues File", filetypes=[("Excel Files", "*.xlsx")])
if not hcl_df_path:
    print("HCL queues file not selected. Exiting...")
    exit()

hcl_df = pd.read_excel(hcl_df_path)

def extract_date_year(opened_date):
    opened_date_str = opened_date.strftime("%m/%d/%y %H:%M")
    date_obj = datetime.strptime(opened_date_str, "%m/%d/%y %H:%M")
    return date_obj.month, date_obj.year

def extract_date_year1(opened_date):
    opened_date_str = opened_date.strftime("%m/%d/%y %H:%M")
    date_obj = datetime.strptime(opened_date_str, "%m/%d/%y %H:%M")
    return f'{date_obj.year}-{date_obj.month}'

def Reopen_column(reopen_count):
    if pd.notna(reopen_count) and reopen_count >=1:
        return 1
    else:
        return 0

def days_between_dates(date1, date2):
    format = '%Y-%m-%d %H:%M:%S' 
    date1_obj = datetime.strptime(date1, format)
    date2_obj = datetime.strptime(date2, format)
    time_difference = date2_obj - date1_obj
    days = time_difference.days
    return days

def Daysold_column(created_date):
    created_date = created_date.strftime('%Y-%m-%d %H:%M:%S')
    date1 = created_date 
    date2 = datetime.now().strftime('%Y-%m-%d %H:%M:%S') 
    return days_between_dates(date1, date2 )


def Olderthan30days_column(days_old):
    if pd.notna(days_old) and days_old >=30:
        return 1
    else:
        return 0

def over_3_hops_column(group_hop_count):
    return 1 if group_hop_count > 3 else 0

def hops_4_6_column(group_hop_count):
    return 1 if 4 <= group_hop_count <= 6 else 0

def hops_6_10_column(group_hop_count):
    return 1 if 6 < group_hop_count <= 10 else 0

def hops_gt_10_column(group_hop_count):
    return 1 if group_hop_count > 10 else 0

def between_7_29_column(opened_date):
    result = Daysold_column(opened_date)
    if result >7 and result<=29:
        return 1
    else:
        return 0
   
def aged_7_10_column(opened_date):
    result = Daysold_column(opened_date)
    if result >=7 and result<=10:
        return 1
    else:
        return 0
   
def aged_11_20_column(opened_date):
    result = Daysold_column(opened_date)
    if result >10 and result<=20:
        return 1
    else:
        return 0
   
def aged_21_29_column(opened_date):
    result = Daysold_column(opened_date)
    if result >20 and result<=29:
        return 1
    else:
        return 0

def ticketunassignedover1day_column(assigned_to):
    if pd.notna(assigned_to):
        return 0
    else:
        return 1

def create_work_columns(assignment_group):   
    match = hcl_df[hcl_df["Name"] == assignment_group]
    if not match.empty:
        HCLQueue = match.iloc[0]["Name"]
        Tower = match.iloc[0]["Tower"]
        SubTower = match.iloc[0]["Sub-Tower"]
        SupportOrganization = match.iloc[0]["Support Organization"]
        return HCLQueue, Tower, SubTower, SupportOrganization
    else:
        return "NO", "NO", "NO", "NO"

def calculate_p2(row):
    if "Priority 2" in row["Priority"]:
        if row["Hops 6-10"] + row["Between 7-29 days"] == 2:
            return 1
        else:
            return 0
    else:
        return 0
   
def calculate_risk_rating(row):
    if row['Priority'] == "Priority 1 - Critical":
        return "P1 - Not User Champions"
    elif row['Priority'] == "Priority 2 - High":
        return "P2 - Not User Champions"
    elif row['Older than 30 days'] >= 1:
        return "Over 30 - SPOC Responsibility"
    else:
        weight = {
        "Reopen": 15,
        "Aged 7-10": 5,
        "Aged 11-20": 10,
        "Aged 21-29": 20,
        "Ticket Unassigned over 1 day": 5,
        "Hop Count 4 - 6": 15,
        "Hope Count 6 - 10": 25,
        "Hops >10": 50
        }
        total_weight = row['Reopen'] * weight['Reopen'] + row['Aged 7-10'] * weight['Aged 7-10'] + row['Aged 11-20'] * weight['Aged 11-20'] + row['Aged 21-29'] * weight['Aged 21-29'] + row['Ticket Unassigned over 1 day'] * weight['Ticket Unassigned over 1 day'] + row['Hops 4-6'] * weight['Hop Count 4 - 6'] + row['Hops 6-10'] * weight['Hope Count 6 - 10'] + row['Hops>10'] * weight['Hops >10']
        return total_weight
    
def calculate_focus_level(row):
    if row['Risk Rating'] == "P1 - Not User Champions":
        return "Do Not Engage"
    elif row['Risk Rating'] == "P2 - Not User Champions":
        return "Do Not Engage"
    elif row['Older than 30 days'] >= 1:
        return "SPOC"
    elif row['Risk Rating'] >= 46:
        return "1"
    elif 5 <= row['Risk Rating'] <= 15:
        return "4"
    elif 16 <= row['Risk Rating'] <= 25:
        return "3"
    elif 26 <= row['Risk Rating'] <= 45:
        return "2"
    else:
        return "0"
    
def update_spoc_columns(row):
    hcl_queue = row['HCL Queue?']
    assignment_group = row['Assignment Group']
    spoc_name = None
    spoc_email_address = None
    support_group_manager_email = None

    if hcl_queue == "NO":
        df.at[row.name, 'SPOC'] = "Not an HCL Queue"
        df.at[row.name, 'SPOC Email'] = "Not an HCL Queue"
    elif hcl_queue in hcl_df['Name'].values:
       
        spoc_name = hcl_df.loc[hcl_df['Name'] == hcl_queue, 'SPOC Name'].iloc[0]
        spoc_email_address = hcl_df.loc[hcl_df['Name'] == hcl_queue, 'SPOC Email Address'].iloc[0]
        support_group_manager_email = hcl_df.loc[hcl_df['Name'] == hcl_queue, 'Support Group Manager Email'].iloc[0]

    df.at[row.name, 'SPOC'] = "SGM" if spoc_name == 0 else spoc_name
    df.at[row.name, 'SPOC Email'] = support_group_manager_email if spoc_name == 0 else spoc_email_address

    df.at[row.name, 'SDM Name'] = hcl_df.loc[hcl_df['Name'] == assignment_group, 'Support Group Manager'].iloc[0] if assignment_group in hcl_df['Name'].values else "none listed"
    df.at[row.name, 'SDM Email'] = hcl_df.loc[hcl_df['Name'] == assignment_group, 'Support Group Manager Email'].iloc[0] if assignment_group in hcl_df['Name'].values else "none listed"


def set_cell_color(ws, row, color):
    for cell in ws[row]:
        cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")


if os.path.exists(output_file_path):
    os.remove(output_file_path)


df = pd.read_excel(input_file_path)

df.rename(columns={
    'HCL reference number': 'Number',
    'Caller': 'Requester',
    'Configuration item': 'Affected CI/Service',
    'Assigned to': 'Assigned To',
    'Assignment group':'Assignment Group',
    'Reassignment count':'Group Hop count',
    'State':'Status',
    

}, inplace=True)

# Print new column names
print("New column names:")
print(df.columns.tolist())


df['Month'], df['Year'] = zip(*df['Opened'].apply(extract_date_year))
df['HCL Queue?'], df['Tower'], df['Sub Tower'], df['Support Organization'] = zip(*df['Assignment Group'].apply(create_work_columns))
df['Reopen'] = df['Reopen count'].apply(Reopen_column)
df['Days old'] = df['Created'].apply(Daysold_column)
df['Older than 30 days'] = df['Days old'].apply(Olderthan30days_column)
df['P2?'] = ''
df['Over 3 Hops'] = df['Group Hop count'].apply(over_3_hops_column)
df['Hops 4-6'] = df['Group Hop count'].apply(hops_4_6_column)
df['Hops 6-10'] = df['Group Hop count'].apply(hops_6_10_column)
df['Hops>10'] = df['Group Hop count'].apply(hops_gt_10_column)
df['Between 7-29 days'] = df['Opened'].apply(between_7_29_column)
df['Aged 7-10'] = df['Opened'].apply(aged_7_10_column)
df['Aged 11-20'] = df['Opened'].apply(aged_11_20_column)
df['Aged 21-29'] = df['Opened'].apply(aged_21_29_column)
df['Ticket Unassigned over 1 day'] = df['Assigned To'].apply(ticketunassignedover1day_column)
df['P2?'] = df.apply(calculate_p2, axis=1)
df['Risk Rating'] = df.apply(calculate_risk_rating, axis=1) 
df['Focus Level'] = df.apply(calculate_focus_level, axis=1)
df['Year-Month'] = df['Opened'].apply(extract_date_year1)
df['SPOC'] = "Not an Hcl Queue"
df['SPOC Email'] = "Not an Hcl Queue"
df['SDM Name'] = "none listed"
df['SDM Email'] = "none listed" 

df['Sub Tower'] = df['Sub Tower'].fillna(0)

df.apply(update_spoc_columns, axis=1)



df.to_excel(output_file_path, index=False)
wb = load_workbook(output_file_path)
ws = wb.active
set_cell_color(ws, "1", 'F5D562')
wb.save(output_file_path)
print("Completed")